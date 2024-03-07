from abc import ABC, abstractmethod
import os
import pandas as pd
from openpyxl import load_workbook
import sqlite_utils
from enum import Enum


class Data(ABC):
    def __init__(self):
        super().__init__()

    @staticmethod
    @abstractmethod
    def download():
        """Abstract static download method. Subclasses should implement this."""
        pass

    @staticmethod
    @abstractmethod
    def clean():
        """Abstract static clean method. Subclasses should implement this."""
        pass

class OfficeType(Enum):
    MAYOR = "Mayor"
    COUNCILLOR = "City Councillor"
    TDSB_TRUSTEE = "TDSB Trustee"
    TCDSB_TRUSTEE = "TCDSB Trustee"

class ContributionSearchResults(Data):
    raw_data_dir = os.path.join(os.getcwd(), "raw_data/contribution_search_results/2022/")

    @classmethod
    def download(cls):
        pass

    @classmethod
    def clean(cls):
        file = os.path.join(cls.raw_data_dir,"contribution-search-results.xls")
        df = pd.read_excel(file,usecols="A:L", skiprows=range(7))
        df.columns = (
            df.columns.str.replace("\n", " ", regex=False)
            .str.replace("\s+", " ", regex=True)
            .str.replace(" */ *", "/", regex=True)
            .str.strip()
        )

        df['Amount'] = df['Amount'].astype(float)
        df['Amount Returned'] = df['Amount Returned'].astype(float)
        df['Contribution Type'] = pd.Categorical(df['Contribution Type'], categories=["Monetary", "Goods/Services"])
        df['Description of Goods/Services'] = df['Description of Goods/Services'].astype('string')
        df['Contributor Type'] = pd.Categorical(df['Contributor Type'], categories=["Candidate", "Individual", "Candidate Spouse", "Third Party Registrant"], ordered=False)
        df['Date Contribution Received'] = pd.to_datetime(df['Date Contribution Received'], format='%b %d, %Y').dt.date
        df['Registered for'] = pd.Categorical(df['Registered for'], categories=["Mayor", "Councillor", "Toronto District School Board", "Toronto Catholic District School Board", "Third Party Advertiser"], ordered=False)
        df['Ward'] = pd.Categorical(df['Ward'], categories=range(26))
        df['Registrant Type'] = pd.Categorical(df["Registrant Type"],categories=["Corporation","Individual"])

        return df

class ElectionsOfficialResults(Data):
    raw_data_dir = os.path.join(os.getcwd(), "raw_data/elections_official_results/2022/")

    @classmethod
    def download(cls):
        pass

    @staticmethod
    def extract_city_ward_results_2022(df):
        return df

    @staticmethod
    def split_df_on_subdivision(df):
        # Find the index positions where 'Subdivision' occurs in the first column
        subdivision_indices = df.index[df.iloc[:, 0] == 'Subdivision'].tolist()

        # Add the last index to cover the final slice
        subdivision_indices.append(df.shape[0])

        # Split the DataFrame on each index found
        split_dfs = [df.iloc[subdivision_indices[n]:subdivision_indices[n+1]] for n in range(len(subdivision_indices)-1)]

        # Optional: reset index for each smaller DataFrame
        #split_dfs = [small_df.reset_index(drop=True) for small_df in split_dfs]

        return split_dfs
    
    @staticmethod
    def clean_election_search_results_2022(file_path,office_type):
        # Initialize an empty list to hold DataFrames
        dfs = []

        # Load the workbook and sheet names using openpyxl
        workbook = load_workbook(filename=file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            # Access the sheet
            sheet = workbook[sheet_name]

            office = sheet['A1'].value

            df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=0, engine="openpyxl")

            # Drop the 2nd row with office
            if office_type == OfficeType.MAYOR or office_type == OfficeType.COUNCILLOR:
                df = pd.concat([df.iloc[:1], df.iloc[2:]]).reset_index(drop=True)


            for city_ward_results in ElectionsOfficialResults.split_df_on_subdivision(df):
                df = city_ward_results
                
                ward = df.iloc[-1, 0]

                # Set first row as column name and delete first row
                df.columns = df.iloc[0]
                df = df[1:]
                

                # Remove trailing blank rows
                while df.tail(1).isna().all(axis=1).values[0]:
                    df = df.iloc[:-1]
                
                ward = df.iloc[-1, 0]

                # Drop the last row (which may contain total values)
                df = df.iloc[:-1]          


                #Not all city wards have the same number of subdvisions. Remove columns with name NaN
                df = df.loc[:, df.columns.notna()]

                # Drop the last column (which may contain total values)
                df = df.iloc[:, :-1]
                
                # Rename the columns and add 'Ward' and 'Ward Name' columns
                df.columns = ['Candidate'] + list(int(c) for c in  df.columns[1:])
            
                df['Ward'] = " ".join(ward.split()[:-1])

                if office_type == OfficeType.MAYOR:
                    df['Office'] = "Mayor"
                elif office_type == OfficeType.COUNCILLOR:
                    ward_number = ward.split()[2]
                    df['Office'] = f"Councillor Ward {ward_number}"
                elif office_type == OfficeType.TDSB_TRUSTEE:
                    ward_number = ward.split()[-2]
                    df['Office'] = f"TDSB Trustee {ward_number}"
                elif office_type == OfficeType.TCDSB_TRUSTEE:
                    ward_number = ward.split()[-2]
                    df['Office'] = f"TCDSB Trustee {ward_number}"

                # Melt the DataFrame to convert it from wide to long format
                df = pd.melt(df, id_vars=['Office', 'Ward', 'Candidate'], var_name='Subdivision', value_name='Vote Count')

                # Append the melted DataFrame to the list
                dfs.append(df)

        # Concatenate all DataFrames in the list into a single DataFrame
        data = pd.concat(dfs, ignore_index=True)

        # Convert data types
        data['Subdivision'] = data['Subdivision'].astype(int)
        data['Vote Count'] = data['Vote Count'].astype(pd.Int64Dtype())

        # Reorder the columns
        data = data[['Office', 'Candidate', 'Ward', 'Subdivision', 'Vote Count']]

        return data


    def clean(cls):
        mayor = cls.clean_election_search_results_2022("raw_data/elections_official_results/2022/2022_Toronto_Poll_By_Poll_Mayor.xlsx",office_type=OfficeType.MAYOR)
        councillor = cls.clean_election_search_results_2022("raw_data/elections_official_results/2022/2022_Toronto_Poll_By_Poll_Councillor.xlsx",office_type=OfficeType.COUNCILLOR)
        tdsb_trustee = cls.clean_election_search_results_2022("raw_data/elections_official_results/2022/2022_Toronto_Poll_By_Poll_Toronto_District_School_Board.xlsx",office_type=OfficeType.TDSB_TRUSTEE)
        tcdsb_trustee = cls.clean_election_search_results_2022("raw_data/elections_official_results/2022/2022_Toronto_Poll_By_Poll_Toronto_Catholic_District_School_Board.xlsx",office_type=OfficeType.TCDSB_TRUSTEE)

        return pd.concat([mayor,councillor,tdsb_trustee,tcdsb_trustee],ignore_index=True)

def build_db():
    if os.path.exists("database.db"):
        os.remove("database.db")

    ContributionSearchResults().clean().to_sql(name="contribution_search_results",con="sqlite:///database.db",index=False)
    ElectionsOfficialResults().clean().to_sql(name="elections_official_results",con="sqlite:///database.db",index=False)

    db = sqlite_utils.Database("database.db")
    db.create_view("[election_winners]", """ 
                                        SELECT 
                                            Candidate, 
                                            Office, 
                                            "Vote Count" 
                                        FROM 
                                            (
                                                SELECT 
                                                    Candidate, 
                                                    Office, 
                                                    SUM("Vote Count") AS "Vote Count", 
                                                    ROW_NUMBER() OVER (
                                                        PARTITION BY Office 
                                                        ORDER BY SUM("Vote Count") DESC
                                                    ) AS rn 
                                                FROM 
                                                    elections_official_results 
                                                GROUP BY 
                                                    Candidate, 
                                                    Office 
                                            ) AS subquery 
                                        WHERE 
                                            rn = 1 
                                        ORDER BY 
                                            CASE 
                                                WHEN Office = 'Mayor' THEN 1 
                                                WHEN Office LIKE 'Councillor Ward%' THEN 2 
                                                WHEN Office LIKE 'TCDSB Trustee Ward%' THEN 3 
                                                WHEN Office LIKE 'TDSB Trustee Ward%' THEN 4 
                                                ELSE 5 
                                            END, 
                                            CASE 
                                                WHEN Office LIKE 'Councillor Ward%' THEN CAST(SUBSTR(Office, 17) AS INTEGER) 
                                                WHEN Office LIKE 'TCDSB Trustee Ward%' THEN CAST(SUBSTR(Office, 20) AS INTEGER) 
                                                WHEN Office LIKE 'TDSB Trustee Ward%' THEN CAST(SUBSTR(Office, 18) AS INTEGER) 
                                                ELSE 0 
                                            END, 
                                            Office;
                                        """)

build_db()